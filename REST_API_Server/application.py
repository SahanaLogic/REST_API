import os
from flask import Flask, jsonify, request
import openpyxl

app = Flask(__name__)

# Path to the Excel database file (same folder as this script)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "drink_database.xlsx")
SHEET_NAME = "Drinks"
HEADERS = ["id", "name", "description"]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
def _get_or_create_workbook():
    """Load the workbook if it exists, otherwise create it with headers."""
    if os.path.exists(DB_PATH):
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)  # write header row on first creation
        wb.save(DB_PATH)
    return wb, ws


def _next_id(ws) -> int:
    """Return the next auto-incremented id (max existing + 1, or 1 if empty)."""
    ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    numeric_ids = [i for i in ids if isinstance(i, int)]
    return max(numeric_ids, default=0) + 1


# ---------------------------------------------------------------------------
# Public database functions
# ---------------------------------------------------------------------------
def populate_list(name: str, description: str) -> dict:
    """Append a new drink to the Excel database and return it as a dict."""
    wb, ws = _get_or_create_workbook()
    new_id = _next_id(ws)
    ws.append([new_id, name, description])
    wb.save(DB_PATH)
    return {"id": new_id, "name": name, "description": description}

def delete_entry(drink_id: int) -> dict | None:
    """Find the row with the given id and delete the entire row.
    Returns the deleted drink as a dict, or None if not found.
    """
    if not os.path.exists(DB_PATH):
        return None
    wb, ws = _get_or_create_workbook()

    for row in ws.iter_rows(min_row=2):
        row_id = row[0].value
        if row_id == drink_id:
            deleted = {
                "id": row[0].value,
                "name": row[1].value,
                "description": row[2].value,
            }
            ws.delete_rows(row[0].row)   # delete the entire row by its row number
            wb.save(DB_PATH)
            return deleted

    return None  # id not found


def retrieve_all_data() -> list[dict]:
    """Parse the Excel file and return all drinks as a list of dicts."""
    if not os.path.exists(DB_PATH):
        return []
    wb, ws = _get_or_create_workbook()
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        drink_id, name, description = row
        if name is not None:  # skip completely empty rows
            result.append({"name": name, "description": description})
    return result


def retrieve_by_id(drink_id: int) -> dict | None:
    """Return the drink with the given id as a dict, or None if not found."""
    if not os.path.exists(DB_PATH):
        return None
    wb, ws = _get_or_create_workbook()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_id, name, description = row
        if row_id == drink_id:
            return {"id": row_id, "name": name, "description": description}
    return None


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return "Drink API — Excel backend"


@app.route("/drinks")
def get_drinks():
    """Return all drinks from the Excel database."""
    return jsonify({"drinks": retrieve_all_data()})


@app.route("/drinks/<int:drink_id>")
def get_drink(drink_id: int):
    """Return a single drink by id."""
    drink = retrieve_by_id(drink_id)
    if drink is None:
        return jsonify({"error": f"Drink with id {drink_id} not found"}), 404
    return jsonify(drink)


@app.route("/drinks/add/<name>/<description>")
def add_drink(name: str, description: str):
    """Quick test route: add a drink via URL params."""
    drink = populate_list(name, description)
    return jsonify({"created": drink}), 201

@app.route("/drinks", methods=['POST'])
def add_drink_by_post():
    drink_name = request.json['name']
    drink_description = request.json['description']
    new_drink = populate_list(drink_name, drink_description)
    return jsonify({'id': f"new drink created with id {new_drink['id']}"}), 201

@app.route("/drinks/<int:drink_id>", methods=['DELETE'])
def delete_drink(drink_id: int):
    """Delete a drink by id."""
    deleted = delete_entry(drink_id)
    if deleted is None:
        return jsonify({"error": f"Drink with id {drink_id} not found"}), 404
    return jsonify({"deleted": deleted}), 200